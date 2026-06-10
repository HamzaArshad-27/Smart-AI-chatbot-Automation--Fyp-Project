from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from django.http import HttpResponse
from apps.orders.models import Order, OrderItem
from apps.products.models import Product, Category
from apps.accounts.models import User
from apps.core.currency import convert_currency
import csv
import openpyxl
from datetime import datetime, timedelta

@login_required
def sales_report(request):
    """Sales report"""
    # Get date range from request
    date_from = request.GET.get('from', (datetime.now() - timedelta(days=30)).date())
    date_to = request.GET.get('to', datetime.now().date())
    
    # Filter orders based on user role
    if request.user.is_admin:
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
    elif request.user.is_company:
        orders = Order.objects.filter(
            company=request.user.company_profile,
            created_at__date__range=[date_from, date_to]
        )
    else:
        orders = Order.objects.filter(user=request.user, created_at__date__range=[date_from, date_to])
    
    # Sales data
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    average_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Daily sales
    daily_sales_raw = orders.extra(
        {'day': "date(created_at)"}
    ).values('day').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('day')
    
    daily_sales = []
    for d in daily_sales_raw:
        total = d['total'] or 0
        count = d['count'] or 1
        d['average'] = float(total) / count
        daily_sales.append(d)
    
    context = {
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'average_order_value': average_order_value,
        'daily_sales': daily_sales,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'reports/sales.html', context)

@login_required
def products_report(request):
    """Products performance report"""
    if request.user.is_company:
        products = Product.objects.filter(company=request.user.company_profile)
    elif request.user.is_seller:
        products = Product.objects.filter(seller=request.user.seller_profile)
    else:
        products = Product.objects.all()
    
    # Product performance
    product_performance = products.annotate(
        sales_count=Sum('order_items__quantity'),
        total_revenue=Sum('order_items__total')
    ).order_by('-total_revenue')[:50]
    
    context = {
        'products': product_performance,
    }
    
    return render(request, 'reports/products.html', context)

@login_required
@user_passes_test(lambda u: u.is_admin)
def users_report(request):
    """Users report"""
    users = User.objects.all().annotate(
        total_orders=Count('orders'),
        total_spent=Sum('orders__total_amount', filter=Q(orders__status='delivered'))
    ).order_by('-total_spent')
    
    context = {
        'users': users,
        'total_users': users.count(),
    }
    
    return render(request, 'reports/users.html', context)




# Additional helper functions for data processing can be added here

@login_required
def dashboard_report(request):
    """Executive dashboard report"""
    # Get current month and previous month data
    today = timezone.now()
    current_month_start = today.replace(day=1, hour=0, minute=0, second=0)
    previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    
    # Current month stats
    current_month_orders = Order.objects.filter(
        created_at__gte=current_month_start,
        status='delivered'
    )
    current_month_revenue = current_month_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Previous month stats
    previous_month_orders = Order.objects.filter(
        created_at__gte=previous_month_start,
        created_at__lt=current_month_start,
        status='delivered'
    )
    previous_month_revenue = previous_month_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Calculate growth
    revenue_growth = ((current_month_revenue - previous_month_revenue) / previous_month_revenue * 100) if previous_month_revenue > 0 else 0
    
    # Top products
    top_products = OrderItem.objects.values('product__name', 'product__id').annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total')
    ).order_by('-total_revenue')[:10]
    
    # Recent activity
    recent_orders = Order.objects.select_related('user', 'company').order_by('-created_at')[:10]
    
    context = {
        'current_month_revenue': current_month_revenue,
        'previous_month_revenue': previous_month_revenue,
        'revenue_growth': revenue_growth,
        'current_month_orders': current_month_orders.count(),
        'previous_month_orders': previous_month_orders.count(),
        'top_products': top_products,
        'recent_orders': recent_orders,
    }
    
    return render(request, 'reports/dashboard.html', context)

@login_required
def category_report(request):
    """Category performance report"""
    categories = Category.objects.annotate(
        total_products=Count('products', filter=Q(products__is_active=True)),
        total_sold=Sum('products__order_items__quantity'),
        total_revenue=Sum('products__order_items__total')
    ).order_by('-total_revenue')
    
    context = {
        'categories': categories,
        'total_revenue': categories.aggregate(total=Sum('total_revenue'))['total'] or 0,
    }
    
    return render(request, 'reports/categories.html', context)

@login_required
def seller_report(request):
    """Seller performance report"""
    if request.user.is_company:
        sellers = User.objects.filter(
            role='seller',
            seller_profile__company=request.user.company_profile
        ).annotate(
            total_orders=Count('seller_profile__order_items__order', distinct=True),
            total_revenue=Sum('seller_profile__order_items__total'),
            total_products=Count('seller_profile__products', distinct=True)
        )
    else:
        sellers = User.objects.filter(role='seller').annotate(
            total_orders=Count('seller_profile__order_items__order', distinct=True),
            total_revenue=Sum('seller_profile__order_items__total'),
            total_products=Count('seller_profile__products', distinct=True)
        )
    
    context = {
        'sellers': sellers.order_by('-total_revenue'),
        'total_sellers': sellers.count(),
    }
    
    return render(request, 'reports/sellers.html', context)

@login_required
def inventory_report(request):
    """Inventory/Stock report"""
    if request.user.is_company:
        products = Product.objects.filter(company=request.user.company_profile)
    else:
        products = Product.objects.all()
    
    # Low stock products
    low_stock = products.filter(stock_quantity__lte=10, is_active=True)
    
    # Out of stock products
    out_of_stock = products.filter(stock_quantity=0, is_active=True)
    
    # Total inventory value
    inventory_value = products.aggregate(total=Sum('price'))['total'] or 0
    
    context = {
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        'inventory_value': inventory_value,
        'total_products': products.count(),
        'total_stock': products.aggregate(total=Sum('stock_quantity'))['total'] or 0,
    }
    
    return render(request, 'reports/inventory.html', context)

@login_required
def download_report(request):
    """Enhanced download report with multiple formats"""
    report_type = request.GET.get('type', 'sales')
    format_type = request.GET.get('format', 'csv')
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    
    currency_code = request.session.get('currency', 'USD')
    from apps.core.currency import CURRENCY_SYMBOLS
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')
    
    # Get data based on report type
    if report_type == 'sales':
        data = get_sales_data(date_from, date_to, currency_code, currency_symbol)
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d')}"
    elif report_type == 'products':
        data = get_products_data(currency_code, currency_symbol)
        filename = f"products_report_{datetime.now().strftime('%Y%m%d')}"
    elif report_type == 'users':
        data = get_users_data(currency_code, currency_symbol)
        filename = f"users_report_{datetime.now().strftime('%Y%m%d')}"
    else:
        data = []
        filename = "report"
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        for row in data:
            writer.writerow(row)
            
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()
        
        for row in data:
            ws.append(row)
        
        # Style the header row
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF")
        
        wb.save(response)
    
    return response

def get_sales_data(date_from=None, date_to=None, currency_code='USD', currency_symbol='$'):
    """Helper function to get sales data for export"""
    orders = Order.objects.filter(status='delivered')
    
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)
    
    data = [['Date', 'Order #', 'Customer', 'Amount', 'Status']]
    for order in orders:
        converted_amount = convert_currency(order.total_amount, currency_code, 'USD')
        data.append([
            order.created_at.strftime('%Y-%m-%d'),
            order.order_number,
            order.user.email,
            f"{currency_symbol}{converted_amount:.2f}",
            order.get_status_display()
        ])
    
    return data

def get_products_data(currency_code='USD', currency_symbol='$'):
    """Helper function to get products data for export"""
    products = Product.objects.filter(is_active=True)
    
    data = [['Product Name', 'SKU', 'Category', 'Price', 'Stock', 'Total Sold', 'Revenue']]
    for product in products:
        total_sold = product.order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        revenue = product.order_items.aggregate(Sum('total'))['total__sum'] or 0
        
        converted_price = convert_currency(product.price, currency_code, 'USD')
        converted_revenue = convert_currency(revenue, currency_code, 'USD')
        data.append([
            product.name,
            product.sku or 'N/A',
            product.category.name if product.category else 'Uncategorized',
            f"{currency_symbol}{converted_price:.2f}",
            product.stock_quantity,
            total_sold,
            f"{currency_symbol}{converted_revenue:.2f}"
        ])
    
    return data

def get_users_data(currency_code='USD', currency_symbol='$'):
    """Helper function to get users data for export"""
    users = User.objects.all()
    
    data = [['Email', 'Role', 'Phone', 'Status', 'Joined Date', 'Total Orders', 'Total Spent']]
    for user in users:
        total_orders = user.orders.filter(status='delivered').count()
        total_spent = user.orders.filter(status='delivered').aggregate(Sum('total_amount'))['total__sum'] or 0
        
        converted_spent = convert_currency(total_spent, currency_code, 'USD')
        data.append([
            user.email,
            user.get_role_display(),
            user.phone or 'N/A',
            'Active' if user.is_active else 'Inactive',
            user.date_joined.strftime('%Y-%m-%d'),
            total_orders,
            f"{currency_symbol}{converted_spent:.2f}"
        ])
    
    return data

@login_required
def sales_api(request):
    """JSON API for real-time sales reporting"""
    from django.http import JsonResponse
    
    date_from_str = request.GET.get('from')
    date_to_str = request.GET.get('to')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = (datetime.now() - timedelta(days=30)).date()
    else:
        date_from = (datetime.now() - timedelta(days=30)).date()
        
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
        
    # Filter orders based on user role
    if request.user.is_admin:
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
    elif request.user.role == 'company':
        orders = Order.objects.filter(
            company=request.user.company_profile,
            created_at__date__range=[date_from, date_to]
        )
    else:
        orders = Order.objects.filter(user=request.user, created_at__date__range=[date_from, date_to])
        
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    average_order_value = float(total_revenue / total_orders) if total_orders > 0 else 0
    
    currency_code = request.session.get('currency', 'USD')
    from apps.core.currency import CURRENCY_SYMBOLS
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, '$')
    
    total_revenue_converted = convert_currency(total_revenue, currency_code, 'USD')
    average_order_value_converted = convert_currency(average_order_value, currency_code, 'USD')
    
    # Status distribution
    status_counts = list(orders.values('status').annotate(count=Count('id')))
    # Payment method distribution
    method_counts = list(orders.values('payment_method').annotate(count=Count('id')))
    # Payment status distribution
    pstatus_counts = list(orders.values('payment_status').annotate(count=Count('id')))
    
    # Daily sales
    daily_sales_qs = orders.extra(
        select={'day': "date(created_at)"}
    ).values('day').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('day')
    
    daily_data = []
    for item in daily_sales_qs:
        day_val = item['day']
        if day_val:
            if isinstance(day_val, str):
                day_str = day_val
            else:
                day_str = day_val.strftime('%Y-%m-%d')
            daily_data.append({
                'day': day_str,
                'total': float(convert_currency(item['total'] or 0, currency_code, 'USD')),
                'count': item['count']
            })
            
    # Format status labels and values
    status_data = {item['status']: item['count'] for item in status_counts}
    method_data = {item['payment_method']: item['count'] for item in method_counts}
    pstatus_data = {item['payment_status']: item['count'] for item in pstatus_counts}
    
    return JsonResponse({
        'success': True,
        'currency_code': currency_code,
        'currency_symbol': currency_symbol,
        'summary': {
            'total_orders': total_orders,
            'total_revenue': float(total_revenue_converted),
            'average_order_value': float(average_order_value_converted),
        },
        'daily_sales': daily_data,
        'status_distribution': status_data,
        'method_distribution': method_data,
        'payment_status_distribution': pstatus_data
    })


@login_required
def admin_reports(request):
    if not (request.user.role == 'admin' or request.user.is_superuser):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You are not authorized to view admin reports.")
    
    report_type = request.GET.get('report', 'sales_overview')
    context = {'report_type': report_type}
    
    # Base queries for date range filtering (last 30 days default)
    date_from_str = request.GET.get('from')
    date_to_str = request.GET.get('to')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = (datetime.now() - timedelta(days=30)).date()
    else:
        date_from = (datetime.now() - timedelta(days=30)).date()
        
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
        
    context.update({
        'date_from': date_from,
        'date_to': date_to,
    })

    if report_type == 'sales_overview' or report_type == 'revenue_analytics':
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        avg_value = total_revenue / total_orders if total_orders > 0 else 0
        
        daily_sales = orders.extra(
            select={'day': "date(created_at)"}
        ).values('day').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('day')
        
        context.update({
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'average_order_value': avg_value,
            'daily_sales': daily_sales,
        })
        
    elif report_type == 'order_status':
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
        status_counts = orders.values('status').annotate(count=Count('id')).order_by('-count')
        context.update({
            'status_counts': status_counts,
        })
        
    elif report_type == 'product_performance':
        products_qs = Product.objects.annotate(
            sales_count=Sum('order_items__quantity', filter=Q(order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('order_items__total', filter=Q(order_items__order__created_at__date__range=[date_from, date_to]))
        ).filter(sales_count__gt=0).order_by('-total_revenue')[:50]
        context.update({
            'products': products_qs,
        })
        
    elif report_type == 'customer_growth':
        from django.db.models.functions import TruncMonth
        customers = User.objects.filter(role='customer', date_joined__date__range=[date_from, date_to])
        growth = customers.annotate(
            month=TruncMonth('date_joined')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
        
        context.update({
            'growth': growth,
            'total_new_customers': customers.count(),
        })
        
    elif report_type == 'vendor_performance':
        from apps.companies.models import Company
        companies = Company.objects.annotate(
            total_orders=Count('orders', filter=Q(orders__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('orders__total_amount', filter=Q(orders__created_at__date__range=[date_from, date_to]))
        ).order_by('-total_revenue')
        context.update({
            'companies': companies,
        })
        
    elif report_type == 'inventory_stock':
        products = Product.objects.all().order_by('stock_quantity')
        low_stock_count = products.filter(stock_quantity__lte=5).count()
        out_of_stock_count = products.filter(stock_quantity=0).count()
        context.update({
            'products': products,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
        })
        
    elif report_type == 'refund_return':
        refunded_orders = Order.objects.filter(
            Q(status='refunded') | Q(payment_status='refunded') | Q(status='cancelled'),
            created_at__date__range=[date_from, date_to]
        )
        total_refunded_amount = refunded_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        context.update({
            'refunded_orders': refunded_orders,
            'total_refunded_orders': refunded_orders.count(),
            'total_refunded_amount': total_refunded_amount,
        })
        
    elif report_type == 'payment_analytics':
        orders = Order.objects.filter(created_at__date__range=[date_from, date_to])
        payment_methods = orders.values('payment_method').annotate(
            count=Count('id'),
            total_amount=Sum('total_amount')
        ).order_by('-total_amount')
        context.update({
            'payment_methods': payment_methods,
        })
        
    elif report_type == 'category_performance':
        categories = Category.objects.annotate(
            total_products=Count('products', filter=Q(products__is_active=True)),
            total_sold=Sum('products__order_items__quantity', filter=Q(products__order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('products__order_items__total', filter=Q(products__order_items__order__created_at__date__range=[date_from, date_to]))
        ).order_by('-total_revenue')
        context.update({
            'categories': categories,
        })
        
    elif report_type == 'business_summary':
        from django.db.models.functions import TruncMonth
        summary = Order.objects.filter(
            created_at__date__range=[date_from, date_to]
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total_orders=Count('id'),
            total_revenue=Sum('total_amount')
        ).order_by('-month')
        context.update({
            'summary': summary,
        })
        
    return render(request, 'reports/admin_reports.html', context)


@login_required
def company_reports(request):
    if not (request.user.role in ['company', 'seller']):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You are not authorized to view company reports.")
        
    if request.user.role == 'company':
        company = request.user.company_profile
    else:
        company = request.user.seller_profile.company
        
    report_type = request.GET.get('report', 'vendor_sales')
    context = {'report_type': report_type, 'company': company}
    
    date_from_str = request.GET.get('from')
    date_to_str = request.GET.get('to')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = (datetime.now() - timedelta(days=30)).date()
    else:
        date_from = (datetime.now() - timedelta(days=30)).date()
        
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
        
    context.update({
        'date_from': date_from,
        'date_to': date_to,
    })

    if report_type == 'vendor_sales':
        orders = Order.objects.filter(company=company, created_at__date__range=[date_from, date_to])
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        
        daily_sales = orders.extra(
            select={'day': "date(created_at)"}
        ).values('day').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('day')
        
        context.update({
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'daily_sales': daily_sales,
        })
        
    elif report_type == 'product_sales':
        products = Product.objects.filter(company=company).annotate(
            sales_count=Sum('order_items__quantity', filter=Q(order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('order_items__total', filter=Q(order_items__order__created_at__date__range=[date_from, date_to]))
        ).order_by('-total_revenue')
        context.update({
            'products': products,
        })
        
    elif report_type == 'revenue_profit':
        items = OrderItem.objects.filter(product__company=company, order__created_at__date__range=[date_from, date_to])
        
        total_revenue = 0
        total_cost = 0
        
        for item in items:
            total_revenue += item.total
            total_cost += (item.product.cost_per_item * item.quantity)
            
        estimated_profit = total_revenue - total_cost
        context.update({
            'total_revenue': total_revenue,
            'total_cost': total_cost,
            'estimated_profit': estimated_profit,
            'items_count': items.count(),
        })
        
    elif report_type == 'inventory_mgmt' or report_type == 'low_stock':
        products = Product.objects.filter(company=company).order_by('stock_quantity')
        low_stock = products.filter(stock_quantity__lte=F('low_stock_threshold'))
        context.update({
            'products': products,
            'low_stock': low_stock,
            'low_stock_count': low_stock.count(),
        })
        
    elif report_type == 'order_fulfillment':
        items = OrderItem.objects.filter(product__company=company, order__created_at__date__range=[date_from, date_to])
        status_counts = items.values('status').annotate(count=Count('id')).order_by('-count')
        context.update({
            'status_counts': status_counts,
        })
        
    elif report_type == 'rating_reviews':
        products = Product.objects.filter(company=company).exclude(total_reviews=0).order_by('-average_rating')
        context.update({
            'products': products,
        })
        
    elif report_type == 'customer_purchase':
        orders = Order.objects.filter(company=company, created_at__date__range=[date_from, date_to]).select_related('user')
        customers = {}
        for o in orders:
            email = o.user.email
            if email not in customers:
                customers[email] = {
                    'name': o.user.get_full_name() or email,
                    'email': email,
                    'phone': o.user.phone or 'N/A',
                    'orders_count': 0,
                    'total_spent': 0
                }
            customers[email]['orders_count'] += 1
            customers[email]['total_spent'] += o.total_amount
            
        context.update({
            'customers': sorted(customers.values(), key=lambda c: c['total_spent'], reverse=True),
        })
        
    elif report_type == 'best_selling':
        products = Product.objects.filter(company=company).annotate(
            sales_count=Sum('order_items__quantity', filter=Q(order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('order_items__total', filter=Q(order_items__order__created_at__date__range=[date_from, date_to]))
        ).filter(sales_count__gt=0).order_by('-sales_count')[:15]
        context.update({
            'products': products,
        })
        
    elif report_type == 'return_analysis':
        refunded_items = OrderItem.objects.filter(
            product__company=company,
            status__in=['refunded', 'cancelled'],
            order__created_at__date__range=[date_from, date_to]
        )
        context.update({
            'refunded_items': refunded_items,
            'refunded_count': refunded_items.count(),
        })
        
    elif report_type == 'discount_campaign':
        products = Product.objects.filter(company=company, compare_price__gt=F('price')).annotate(
            sales_count=Sum('order_items__quantity', filter=Q(order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('order_items__total', filter=Q(order_items__order__created_at__date__range=[date_from, date_to]))
        )
        context.update({
            'products': products,
        })
        
    elif report_type == 'seller_earnings':
        sellers = User.objects.filter(
            role='seller',
            seller_profile__company=company
        ).annotate(
            total_orders=Count('seller_profile__order_items__order', distinct=True, filter=Q(seller_profile__order_items__order__created_at__date__range=[date_from, date_to])),
            total_revenue=Sum('seller_profile__order_items__total', filter=Q(seller_profile__order_items__order__created_at__date__range=[date_from, date_to])),
            total_products=Count('seller_profile__products', distinct=True)
        ).order_by('-total_revenue')
        context.update({
            'sellers': sellers,
        })
        
    return render(request, 'reports/company_reports.html', context)


@login_required
def customer_reports(request):
    if not (request.user.role in ['customer', 'retailer']):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You are not authorized to view buyer reports.")
        
    report_type = request.GET.get('report', 'purchase_history')
    context = {'report_type': report_type}
    
    date_from_str = request.GET.get('from')
    date_to_str = request.GET.get('to')
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = (datetime.now() - timedelta(days=30)).date()
    else:
        date_from = (datetime.now() - timedelta(days=30)).date()
        
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = datetime.now().date()
    else:
        date_to = datetime.now().date()
        
    context.update({
        'date_from': date_from,
        'date_to': date_to,
    })

    if report_type == 'purchase_history':
        orders = Order.objects.filter(user=request.user, created_at__date__range=[date_from, date_to])
        context.update({
            'orders': orders,
            'total_orders': orders.count(),
            'total_spent': orders.aggregate(total=Sum('total_amount'))['total'] or 0,
        })
        
    elif report_type == 'order_tracking':
        active_orders = Order.objects.filter(
            user=request.user, 
            status__in=['pending', 'approved', 'processing', 'shipped']
        )
        context.update({
            'orders': active_orders,
        })
        
    elif report_type == 'spending_summary' or report_type == 'spending_analysis':
        from django.db.models.functions import TruncMonth
        spending = Order.objects.filter(
            user=request.user,
            created_at__date__range=[date_from, date_to],
            status='delivered'
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Sum('total_amount'),
            orders_count=Count('id')
        ).order_by('-month')
        
        context.update({
            'spending': spending,
            'total_spent_lifetime': Order.objects.filter(user=request.user, status='delivered').aggregate(total=Sum('total_amount'))['total'] or 0
        })
        
    elif report_type == 'product_interests':
        from apps.ai_assistant.models import UserProductInterest
        interests = UserProductInterest.objects.filter(user=request.user).select_related('product').order_by('-weight')
        context.update({
            'interests': interests,
        })
        
    elif report_type == 'return_refund_history':
        refunded_orders = Order.objects.filter(
            user=request.user,
            status__in=['cancelled', 'refunded'],
            created_at__date__range=[date_from, date_to]
        )
        context.update({
            'orders': refunded_orders,
        })
        
    return render(request, 'reports/customer_reports.html', context)